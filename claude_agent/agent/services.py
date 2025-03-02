import os
import anthropic
from django.conf import settings
import tempfile
import mimetypes
import docx
import PyPDF2
import csv
import json
from openpyxl import load_workbook
import openai
import subprocess
import time
import assemblyai as aai
from tempfile import NamedTemporaryFile

class FileProcessor:
    @staticmethod
    def extract_text_from_file(file_path):
        """Extract text content from different file types"""
        mime_type, _ = mimetypes.guess_type(file_path)
        file_extension = os.path.splitext(file_path)[1].lower()
        
        # Обработка по расширению файла, если mime_type не определен
        if mime_type is None and file_extension:
            if file_extension == '.txt':
                mime_type = 'text/plain'
            elif file_extension == '.pdf':
                mime_type = 'application/pdf'
            elif file_extension == '.docx':
                mime_type = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            elif file_extension == '.xlsx':
                mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            elif file_extension == '.csv':
                mime_type = 'text/csv'
            elif file_extension == '.json':
                mime_type = 'application/json'
            elif file_extension in ['.mp3', '.wav', '.ogg', '.m4a', '.flac']:
                mime_type = 'audio/' + file_extension[1:]
        
        # Определяем, является ли файл аудио
        is_audio = mime_type and mime_type.startswith('audio/')
        is_audio_ext = file_extension in ['.mp3', '.wav', '.ogg', '.m4a', '.flac']
        
        # Обработка аудиофайлов
        if is_audio or is_audio_ext:
            # Проверяем наличие настройки AUDIO_PROCESSING_METHOD
            method = getattr(settings, 'AUDIO_PROCESSING_METHOD', 'assemblyai')
            
            if method == 'assemblyai':
                # Используем AssemblyAI (предпочтительный метод)
                print(f"Обработка аудиофайла через AssemblyAI: {file_path}")
                result = AssemblyAIHandler.process_audio(file_path)
                if "error" in result:
                    return f"Ошибка при обработке аудио: {result['error']}"
                return f"ТРАНСКРИПЦИЯ АУДИО:\n\n{result['transcription']}"
            else:
                # Используем локальные методы транскрибации
                print(f"Обработка аудиофайла через локальный метод: {file_path}")
                return f"ТРАНСКРИПЦИЯ АУДИО:\n\n{AudioTranscriber.transcribe_audio(file_path, method)}"
        
        # Попытка прочитать как текст, если mime_type не определен или это текстовый файл
        if mime_type is None or mime_type == 'text/plain' or file_extension == '.txt':
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read()
            except UnicodeDecodeError:
                # Если не удалось прочитать как UTF-8, попробуем другие кодировки
                try:
                    with open(file_path, 'r', encoding='latin-1') as f:
                        return f.read()
                except Exception as e:
                    return f"Ошибка при чтении текстового файла: {str(e)}"
                
        elif mime_type == 'application/pdf':
            text = ""
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    text += page.extract_text() + "\n"
            return text
            
        elif mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
            doc = docx.Document(file_path)
            return "\n".join([para.text for para in doc.paragraphs])
            
        elif mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            wb = load_workbook(file_path)
            text = ""
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    text += " | ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
            return text
            
        elif mime_type == 'text/csv':
            text = ""
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    text += " | ".join(row) + "\n"
            return text
            
        elif mime_type in ['application/json']:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return json.dumps(data, indent=2)
            
        else:
            return f"Unsupported file format: {mime_type}"

class ClaudeService:
    def __init__(self):
        """Инициализация сервиса с проверкой наличия API-ключа"""
        api_key = settings.CLAUDE_API_KEY
        if not api_key:
            raise ValueError("CLAUDE_API_KEY не найден в настройках. Проверьте файл .env")
        
        self.client = anthropic.Anthropic(
            api_key=api_key,
            # Увеличиваем timeout для больших запросов
            timeout=60.0
        )
        
        # Проверяем модель из настроек или используем стандартную
        self.default_model = getattr(settings, 'MODEL_NAME', 'claude-3-sonnet-20240229')
        # Альтернативные модели для автоматического переключения
        self.fallback_models = [
            'claude-3-haiku-20240307',
            'claude-3-opus-20240229',
            'claude-3-5-sonnet-20240620',
            'claude-instant-1.2',
            'claude-2.0',
        ]
        print(f"Используется модель: {self.default_model}")
    
    def _send_api_request(self, model, system_message, prompt):
        """Отправляет запрос к API с указанной моделью и обрабатывает ошибки"""
        try:
            print(f"Отправка запроса к Claude API с моделью {model}...")
            response = self.client.messages.create(
                model=model,
                max_tokens=4000,
                temperature=0,
                system=system_message,
                messages=[
                    {"role": "user", "content": prompt}
                ]
            )
            print("Ответ успешно получен!")
            return response.content[0].text, None
        except Exception as e:
            return None, e
    
    def compare_documents(self, documents, custom_prompt=None):
        """
        Takes a list of document objects, extracts their content,
        and sends to Claude for comparative analysis
        
        Args:
            documents: QuerySet of Document objects
            custom_prompt: Optional custom prompt provided by user
        """
        document_contents = []
        
        for doc in documents:
            print(f"Обработка документа: {doc.name} (тип: {doc.file_type})")
            with tempfile.NamedTemporaryFile(delete=False) as temp:
                temp.write(doc.file.read())
                temp_path = temp.name
            
            try:
                content = FileProcessor.extract_text_from_file(temp_path)
                file_extension = os.path.splitext(doc.name)[1].lower()
                print(f"Расширение файла: {file_extension}, определенный тип: {doc.file_type}")
                
                # Проверяем, получен ли текст
                if content and not content.startswith("Unsupported file format") and not content.startswith("Ошибка при чтении"):
                    print(f"Успешно извлечен текст из {doc.name}. Размер: {len(content)} символов")
                else:
                    print(f"Проблема с извлечением текста из {doc.name}: {content}")
                
                document_contents.append({
                    "name": doc.name,
                    "type": doc.file_type or f"Файл{file_extension}",
                    "content": content
                })
            finally:
                os.unlink(temp_path)
        
        # Prepare prompt for Claude
        if custom_prompt:
            prompt = self._build_custom_prompt(document_contents, custom_prompt)
            system_message = "You are a helpful assistant. Follow the user's instructions carefully regarding the documents."
        else:
            prompt = self._build_comparison_prompt(document_contents)
            system_message = "You are an expert analyst who performs thorough comparative analysis of documents."
        
        # Сначала пробуем основную модель
        result, error = self._send_api_request(self.default_model, system_message, prompt)
        if result:
            return result
            
        # Если основная модель не сработала, пробуем резервные модели
        print(f"Ошибка при использовании основной модели: {error}")
        print("Пробую альтернативные модели...")
        
        for fallback_model in self.fallback_models:
            print(f"Попытка использовать модель: {fallback_model}")
            result, error = self._send_api_request(fallback_model, system_message, prompt)
            if result:
                print(f"Удалось получить ответ от модели {fallback_model}")
                return result
            print(f"Модель {fallback_model} тоже не работает: {error}")
        
        # Если все модели не сработали, возвращаем сообщение об ошибке
        return f"Не удалось получить ответ ни от одной доступной модели Claude. Проверьте ваш API-ключ и доступ к моделям Claude. Последняя ошибка: {str(error)}"
    
    def _build_comparison_prompt(self, document_contents):
        """Builds a structured prompt for Claude to compare documents"""
        prompt = "Please perform a detailed comparative analysis of the following documents:\n\n"
        
        for i, doc in enumerate(document_contents, 1):
            prompt += f"## DOCUMENT {i}: {doc['name']} (Format: {doc['type']})\n\n"
            prompt += doc['content'][:10000]  # Limiting content length
            prompt += "\n\n---\n\n"
        
        prompt += """
Please analyze these documents and provide:
1. A summary of each document
2. Key similarities between the documents
3. Notable differences between the documents
4. Any insights or patterns you observe
5. A conclusion about how these documents relate to each other

Format your response in a structured, clear manner using markdown.
"""
        return prompt
    
    def _build_custom_prompt(self, document_contents, custom_prompt):
        """
        Builds a custom prompt using the user's instructions
        
        Args:
            document_contents: List of dictionaries with document contents
            custom_prompt: User-provided instructions
        """
        prompt = f"I have the following documents and I need you to: {custom_prompt}\n\n"
        
        for i, doc in enumerate(document_contents, 1):
            prompt += f"## DOCUMENT {i}: {doc['name']} (Format: {doc['type']})\n\n"
            prompt += doc['content'][:10000]  # Limiting content length
            prompt += "\n\n---\n\n"
        
        prompt += """
Please respond to my request based on these documents. 
Format your response in a structured, clear manner using markdown.
"""
        return prompt

class AssemblyAIHandler:
    """Класс для работы с аудиофайлами через AssemblyAI"""
    
    @staticmethod
    def get_api_key():
        api_key = getattr(settings, 'ASSEMBLYAI_API_KEY', None)
        if not api_key:
            raise ValueError("ASSEMBLYAI_API_KEY не настроен в файле .env")
        return api_key
        
    @staticmethod
    def process_audio(file_path, prompt=None):
        """
        Обрабатывает аудиофайл через AssemblyAI и получает анализ от Claude
        
        Args:
            file_path: Путь к аудиофайлу
            prompt: Опциональный запрос для анализа (если не указан, будет выполнено только транскрибирование)
            
        Returns:
            dict: Словарь с результатами транскрибации и анализа
        """
        try:
            # Настройка AssemblyAI API ключа
            aai.settings.api_key = AssemblyAIHandler.get_api_key()
            
            print(f"Начинаем транскрибирование аудиофайла: {file_path}")
            
            # Транскрибация файла
            transcriber = aai.Transcriber()
            transcript = transcriber.transcribe(file_path)
            
            # Результаты
            result = {
                "transcription": transcript.text,
                "analysis": None
            }
            
            # Если указан запрос для анализа, выполняем его
            if prompt:
                print(f"Выполняем анализ контента с запросом: {prompt}")
                # Используем Claude 3.5 Sonnet по умолчанию, но можно настроить на другие модели
                analysis = transcript.lemur.task(
                    prompt,
                    final_model=aai.LemurModel.claude3_5_sonnet
                )
                result["analysis"] = analysis.response
            
            return result
        except Exception as e:
            error_message = f"Ошибка при обработке аудио через AssemblyAI: {str(e)}"
            print(error_message)
            return {"error": error_message}

class AudioTranscriber:
    """Класс для транскрибации аудио-файлов"""
    
    @staticmethod
    def transcribe_with_openai_api(file_path):
        """Транскрибация аудио через OpenAI API"""
        try:
            # Проверяем наличие API ключа OpenAI
            api_key = getattr(settings, 'OPENAI_API_KEY', None)
            if not api_key:
                return "Ошибка: Отсутствует API ключ OpenAI. Добавьте OPENAI_API_KEY в настройки."
            
            client = openai.OpenAI(api_key=api_key)
            
            with open(file_path, "rb") as audio_file:
                transcript = client.audio.transcriptions.create(
                    model="whisper-1", 
                    file=audio_file
                )
            
            return transcript.text
        except Exception as e:
            return f"Ошибка при транскрибации через API OpenAI: {str(e)}"
    
    @staticmethod
    def transcribe_with_local_whisper(file_path):
        """Транскрибация аудио с использованием локальной модели whisper"""
        try:
            # Проверка наличия установленного whisper
            try:
                subprocess.run(["whisper", "--help"], capture_output=True, check=False)
            except FileNotFoundError:
                return "Ошибка: Whisper не установлен. Установите через pip install openai-whisper"
            
            # Запускаем процесс транскрибации
            result = subprocess.run(
                ["whisper", file_path, "--model", "tiny", "--output_format", "txt"],
                capture_output=True,
                text=True
            )
            
            if result.returncode != 0:
                return f"Ошибка при запуске whisper: {result.stderr}"
            
            # Название выходного файла такое же, как входного, но с расширением .txt
            output_file = file_path.rsplit('.', 1)[0] + ".txt"
            
            # Читаем результат транскрипции
            try:
                with open(output_file, 'r', encoding='utf-8') as f:
                    transcription = f.read()
                # Удаляем временный файл
                os.remove(output_file)
                return transcription
            except Exception as file_err:
                return f"Ошибка при чтении файла транскрипции: {str(file_err)}"
        
        except Exception as e:
            return f"Ошибка при локальной транскрибации: {str(e)}"
    
    @staticmethod
    def transcribe_audio(file_path, method="api"):
        """
        Транскрибирует аудио-файл в текст
        
        Args:
            file_path: Путь к аудио-файлу
            method: Метод транскрибации ('api' для OpenAI API или 'local' для локальной модели)
        
        Returns:
            str: Текст транскрипции или сообщение об ошибке
        """
        # Определяем формат аудио-файла
        mime_type, _ = mimetypes.guess_type(file_path)
        file_extension = os.path.splitext(file_path)[1].lower()
        
        # Проверяем, что это аудио-файл
        is_audio = mime_type and mime_type.startswith('audio/')
        is_audio_ext = file_extension in ['.mp3', '.wav', '.ogg', '.m4a', '.flac']
        
        if not (is_audio or is_audio_ext):
            return f"Ошибка: Файл не является аудио-файлом: {mime_type or file_extension}"
        
        # Выбираем метод транскрибации
        if method == "api":
            return AudioTranscriber.transcribe_with_openai_api(file_path)
        elif method == "local":
            return AudioTranscriber.transcribe_with_local_whisper(file_path)
        else:
            return f"Ошибка: Неизвестный метод транскрибации: {method}" 